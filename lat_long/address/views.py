import os
from django.shortcuts import render, redirect
from .forms import UploadFileForm
from openpyxl import load_workbook
import requests
from urllib.parse import urlencode
from django.http import HttpResponse, Http404
from .models import AddressFile
from django.contrib import messages
from django.conf import settings
# Create your views here.


def lat_long(address, datatype="json"):
    endpoint = f"https://maps.googleapis.com/maps/api/geocode/{datatype}"
    key = settings.API_KEY
    params = {"address":address, "key": key}
    url_params = urlencode(params)
    url = f"{endpoint}?{url_params}"
    r = requests.get(url)
    if r.status_code not in range(200,299):
        return {}
    latlng = {}
    try:
        latlng = r.json()['results'][0]['geometry']['location']
    except:
        pass
    return latlng.get('lat'), latlng.get('lng')


def upload(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        print("form: ", request.FILES['file'].file)
        if form.is_valid():
            print("name:",request.FILES['file'] )
            file_name = request.FILES['file']
            if not str(file_name).endswith('xlsx'):
                messages.info(request, 'Incorrect File Format. Please Upload .xlsx file')
                return redirect('/upload')

            wb = load_workbook(filename=request.FILES['file'].file)
            print("wb: ", wb)
            obj = wb.active
            print("obj: ",obj)
            max_row = obj.max_row
            max_col = obj.max_column
            for i in range(1, max_col + 1):
                for j in range(1, max_row + 1):
                    cell_obj = obj.cell(row = j, column = i)
                    if type(cell_obj.value) == str:
                        r = lat_long(cell_obj.value)
                        lat_val = obj.cell(row=cell_obj.row, column=cell_obj.column + 1)
                        lat_val.value = r[0]
                        lng_val = obj.cell(row=cell_obj.row, column=cell_obj.column + 2)
                        lng_val.value = r[1]
                        wb.save(filename=request.FILES['file'].file)
            form.save()
            return redirect('/upload')
    else:
        form = UploadFileForm()

    file = AddressFile.objects.all()
    file_obj = file.last()

    context = {
        'form': form,
        'file_obj': file_obj
    }
    return render(request, 'upload_form.html',context)


def download(request,path):
    file_path = os.path.join(settings.MEDIA_ROOT,path)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/file")
            response['Content-Disposition'] = 'inline;filename='+os.path.basename(file_path)
            return response
    raise Http404
